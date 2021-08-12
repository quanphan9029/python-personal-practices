from re import I
from openpyxl import Workbook
import random



wb = Workbook()
ws = wb.active
ws.title = "Chobai3"

def phep_tinh():
    list_phep_tinh = ['+', '-', 'x', ':']
    count_plus = 0
    count_minus = 0
    count_multiple = 0
    count_divide = 0

    while list_phep_tinh != []:
        for i in range(1, 1001):
            phep_tinh = random.choice(list_phep_tinh)
            cell_value = ws.cell(column=2, row=i)
            if phep_tinh == '+':
                if count_plus == 249:
                    cell_value.value = phep_tinh
                    list_phep_tinh.remove('+')
                else:
                    cell_value.value = phep_tinh
                    count_plus += 1
            elif phep_tinh == '-':
                if count_minus == 249:
                    cell_value.value = phep_tinh
                    list_phep_tinh.remove('-')
                else:
                    cell_value.value = phep_tinh
                    count_minus += 1
            elif phep_tinh == 'x':
                if count_multiple == 249:
                    cell_value.value = phep_tinh
                    list_phep_tinh.remove('x')
                else:
                    cell_value.value = phep_tinh
                    count_multiple += 1
            elif phep_tinh == ':':
                if count_divide == 249:
                    cell_value.value = phep_tinh
                    list_phep_tinh.remove(':')
                else:
                    cell_value.value = phep_tinh
                    count_divide += 1
    print(list_phep_tinh)

def cho_so():
    for i in range(1, 1001):
        if ws.cell(column=2, row=i).value == "+":
            so1 = random.randint(1000, 9999)
            so2 = random.randint(1000, 9999)
            ws.cell(column=1, row=i).value = so1
            ws.cell(column=3, row=i).value = so2
        elif ws.cell(column=2, row=i).value == "-":
            so1 = random.randint(1000, 9999)
            so2 = random.randint(1000, 9999)
            ws.cell(column=1, row=i).value = so1
            ws.cell(column=3, row=i).value = so2
            while so1 < so2 :
                so1 = random.randint(1000, 9999)
                so2 = random.randint(1000, 9999)
                ws.cell(column=1, row=i).value = so1
                ws.cell(column=3, row=i).value = so2
        elif ws.cell(column=2, row=i).value == "x":
            so1 = random.randint(1000, 9999)
            so2 = random.randint(2, 9)
            ws.cell(column=1, row=i).value = so2
            ws.cell(column=3, row=i).value = so1
        elif ws.cell(column=2, row=i).value == ":":
            so1 = random.randint(1000, 9999)
            so2 = random.randint(2, 9)
            ws.cell(column=1, row=i).value = so1
            ws.cell(column=3, row=i).value = so2
            while so1 < so2:
                so1 = random.randint(1000, 9999)
                so2 = random.randint(2, 9)
                ws.cell(column=1, row=i).value = so1
                ws.cell(column=3, row=i).value = so2
        ws.cell(column=4, row=i).value = "="


print(wb.sheetnames)
phep_tinh()
cho_so()

#ten_bai = input(f"Enter your file name: ")
wb.save("Chobai3.xlsx")